from bs4 import BeautifulSoup
import requests
import openpyxl

# Создаем новую книгу Excel и выбираем активный лист
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'Sticker Items'

row = 2 # Строка с которой начнём запись в exel документ

# url для парсинга
url = 'https://csgostash.com/stickers/regular?'
count = 1 # Счётчик страниц

# Пишем скрипт для определения максимального количества страниц
res = requests.get(url) # Делаем запрос
sp = BeautifulSoup(res.content, 'html.parser') # Варим суп

page_list = list() # Создаём список всех страниц

pagination = sp.find('ul', class_='pagination')
pages = pagination.findAll('li')

for i in pages:
    page_list.append(i.text)
max_page = int(page_list[13])


# Параметры для гет запроса
params = {
    'page': count
}
while count < max_page + 1:
    # Гет запрос к странице
    request = requests.get(url, params=params)
    responce = request.content

    # Варим суп
    soup = BeautifulSoup(responce, 'html.parser')

    item_name = [i.text for i in soup.findAll('h3')] # Итератор названий предметов
    item_link = [link['href'] for link in soup.findAll('a', class_='btn btn-default market-button-item')] # Итератор ссылок на прежметы

    for name, link in zip(item_name, item_link):
        sheet.cell(row=row, column=1, value=name)  # Записываем название предмета в первую колонку
        sheet.cell(row=row, column=2, value=link)  # Записываем ссылку на предмет во вторую колонку
        row += 1  # Переходим на следующую строку для следующего предмета

    count += 1 # Счётчик перехода на следующую страницу

workbook.save("sticker_items.xlsx")  # Сохраняем книгу Excel